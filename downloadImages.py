import csv
import urllib2
import os
import sys
import subprocess
from openpyxl import load_workbook

wb = load_workbook('/Users/Pedro/PycharmProjects/BIDHU/docs/galaxies.xlsx');

ws = wb.active

it = -1
cmd = 'java -jar casjobs.jar execute "select distinct ra, dec, petror90_r, specObjId, objId from galaxy WHERE '

for row in ws.iter_rows('I2:I41752'):
    it = it + 1
    if(it > 1000):
        cmd = cmd + ' "'
        proc = subprocess.Popen([cmd, " "], stdout=subprocess.PIPE, shell=True)
        (out, err) = proc.communicate()
        print out

        out = out.split("\n")
        width = 80
        height = 80
        pixelsize = 0.2
        pixelsize2 = 0.2

        for i in range(3, 603):
            aux = out[i].split(",")
            scale = 2 * float(aux[2]) / pixelsize / width
            scale2 = 2 * float(aux[2]) / pixelsize2 / 512
            image = open('/Users/Pedro/PycharmProjects/BIDHU/tests3/galaxy' + str(i - 2) + '.jpg', 'wb')
            image2 = open('/Users/Pedro/PycharmProjects/BIDHU/tests4/galaxy' + str(i - 2) + '.jpg', 'wb')
            image.write(urllib2.urlopen('http://skyservice.pha.jhu.edu/DR12/ImgCutout/getjpeg.aspx?ra=' + str(aux[0]) + '&dec=' + str(aux[1]) + '&width=80&height=80&scale=' + str(scale)).read())
            image2.write(urllib2.urlopen('http://skyservice.pha.jhu.edu/DR12/ImgCutout/getjpeg.aspx?ra=' + str(aux[0]) + '&dec=' + str(aux[1]) + '&width=512&height=512&scale=' + str(scale2)).read())
            image.close()
            image2.close()


        sys.exit()

    if(it > 1):
        cmd = cmd + " OR"
    if it > 0:
        cmd = cmd + ' specObjId = ' + str(row[0].value) + ' ';

